from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from io import BytesIO
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.event import Event
from app.models.event_item import EventItem
from app.models.user import User

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - handled at runtime with a clear error
    load_workbook = None  # type: ignore[assignment]


TARGET_SHEETS: dict[str, int] = {
    "Январь 2026": 1,
    "Февраль 2026": 2,
    "Март 2026": 3,
    "Апрель 2026": 4,
}

SPECIAL_LABELS = {
    "агентские": "agency",
    "ндс": "vat_line",
    "налоги": "taxes",
    "менеджер": "manager_salary",
    "налоги +": "taxes_section",
}


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\xa0", " ").strip()


def label_key(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value).lower()).strip(" :")


def to_decimal(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0.00")
    if isinstance(value, bool):
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if isinstance(value, int | float):
        try:
            return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        except InvalidOperation:
            return Decimal("0.00")
    raw = clean_text(value)
    if not raw:
        return Decimal("0.00")
    raw = raw.replace(" ", "").replace(",", ".")
    raw = re.sub(r"[^0-9.\-]", "", raw)
    if raw in {"", "-", "."}:
        return Decimal("0.00")
    try:
        return Decimal(raw).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return Decimal("0.00")


def money_out(value: Decimal | int | float | None) -> float:
    if value is None:
        return 0.0
    if not isinstance(value, Decimal):
        value = to_decimal(value)
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def parse_any_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = clean_text(value)
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(raw, fmt).date()
        except Exception:
            pass
    return None


def normalize_event_date(raw_value: Any, *, sheet_name: str, sheet_month: int, source_row: int, warnings: list[str]) -> str:
    parsed = parse_any_date(raw_value)
    if not parsed:
        warnings.append(f"{sheet_name} row {source_row}: дата не распознана, поставлено 2026-{sheet_month:02d}-01")
        return date(2026, sheet_month, 1).isoformat()

    day = parsed.day
    # Google Sheets had several wrong years/months in old data. User confirmed
    # the sheet name is the source of truth for January-April 2026.
    try:
        normalized = date(2026, sheet_month, day)
    except ValueError:
        normalized = date(2026, sheet_month, 1)
        warnings.append(f"{sheet_name} row {source_row}: день {day} не подходит месяцу, поставлено {normalized.isoformat()}")
        return normalized.isoformat()

    if parsed.year != 2026 or parsed.month != sheet_month:
        warnings.append(
            f"{sheet_name} row {source_row}: дата {parsed.isoformat()} нормализована в {normalized.isoformat()} по названию листа"
        )
    return normalized.isoformat()


def row_values(row: tuple[Any, ...], max_cols: int = 14) -> list[Any]:
    vals = list(row[:max_cols])
    if len(vals) < max_cols:
        vals.extend([None] * (max_cols - len(vals)))
    return vals


def col_index_by_header(header: list[Any], names: set[str]) -> int | None:
    for idx, value in enumerate(header):
        key = label_key(value)
        if key in names:
            return idx
    return None


def is_event_start(row: list[Any]) -> bool:
    marker = row[0]
    title = clean_text(row[1])
    parsed_date = parse_any_date(row[2])
    if not title or not parsed_date:
        return False
    if isinstance(marker, int | float) and int(marker) == 1:
        return True
    return clean_text(marker) == "1"


def total_row_values(row: list[Any]) -> dict[str, Any]:
    label_values = [label_key(v) for v in row]
    numbers = [to_decimal(v) for v in row]
    budget = Decimal("0.00")
    income = Decimal("0.00")
    status = ""

    for idx, key in enumerate(label_values):
        if key == "общий бюджет" and idx + 1 < len(row):
            budget = to_decimal(row[idx + 1])
        if key == "общий остаток" and idx + 1 < len(row):
            income = to_decimal(row[idx + 1])

    # Newer monthly sheets usually keep final budget in D and final income in I.
    # January old format keeps budget in D and income after "Общий остаток".
    if budget == 0:
        for candidate_idx in (3, 2):
            if candidate_idx < len(row) and to_decimal(row[candidate_idx]) != 0:
                budget = to_decimal(row[candidate_idx])
                break
    if income == 0:
        for candidate_idx in (8, 5, 7):
            if candidate_idx < len(row) and to_decimal(row[candidate_idx]) != 0:
                income = to_decimal(row[candidate_idx])
                break

    for value in row:
        s = label_key(value)
        if s in {"в кассе", "у димы", "у александра", "у саши"}:
            status = clean_text(value)
            break

    return {
        "budget": budget,
        "income": income,
        "cash_status": status,
        "raw": [clean_text(v) if not isinstance(v, (int, float)) else v for v in row[:10]],
    }


@dataclass
class ParsedLine:
    row: int
    name: str
    kind: str
    cost: Decimal = Decimal("0.00")
    fact: Decimal = Decimal("0.00")
    vat: Decimal = Decimal("0.00")
    deduction: Decimal = Decimal("0.00")
    commission: Decimal = Decimal("0.00")

    def as_dict(self) -> dict[str, Any]:
        return {
            "row": self.row,
            "name": self.name,
            "kind": self.kind,
            "cost": money_out(self.cost),
            "fact": money_out(self.fact),
            "vat": money_out(self.vat),
            "deduction": money_out(self.deduction),
            "commission": money_out(self.commission),
        }


@dataclass
class ParsedEvent:
    sheet: str
    source_row: int
    title: str
    raw_date: str
    event_date: str
    old_manager: str
    total_budget: Decimal
    total_income: Decimal
    cash_status: str
    regular_cost_sum: Decimal
    all_cost_sum: Decimal
    fact_sum: Decimal
    vat_sum: Decimal
    deduction_sum: Decimal
    commission_sum: Decimal
    agency_commission: Decimal
    taxes_net: Decimal
    manager_salary: Decimal
    line_count: int
    lines_preview: list[dict[str, Any]]
    lines: list[ParsedLine] = field(default_factory=list, repr=False)

    def as_dict(self) -> dict[str, Any]:
        return {
            "sheet": self.sheet,
            "source_row": self.source_row,
            "title": self.title,
            "raw_date": self.raw_date,
            "event_date": self.event_date,
            "old_manager": self.old_manager,
            "cash_status": self.cash_status,
            "line_count": self.line_count,
            "budget": {
                "total_budget": money_out(self.total_budget),
                "total_income": money_out(self.total_income),
                "regular_cost_sum": money_out(self.regular_cost_sum),
                "all_cost_sum": money_out(self.all_cost_sum),
                "fact_sum": money_out(self.fact_sum),
                "vat_sum": money_out(self.vat_sum),
                "deduction_sum": money_out(self.deduction_sum),
                "commission_sum": money_out(self.commission_sum),
                "agency_commission": money_out(self.agency_commission),
                "taxes_net": money_out(self.taxes_net),
                "manager_salary": money_out(self.manager_salary),
            },
            "lines_preview": self.lines_preview,
        }


def parse_event_block(sheet_name: str, sheet_month: int, rows: list[tuple[int, list[Any]]], warnings: list[str]) -> ParsedEvent | None:
    if not rows:
        return None
    start_row_num, start = rows[0]
    title = clean_text(start[1])
    raw_date = clean_text(start[2])
    event_date = normalize_event_date(start[2], sheet_name=sheet_name, sheet_month=sheet_month, source_row=start_row_num, warnings=warnings)

    header_row = rows[1][1] if len(rows) > 1 else []
    old_manager = clean_text(header_row[1]) if len(header_row) > 1 else ""

    name_col = 1
    cost_col = col_index_by_header(header_row, {"стоимость"})
    fact_col = col_index_by_header(header_row, {"расход", "факт"})
    vat_col = col_index_by_header(header_row, {"ндс"})
    deduction_col = col_index_by_header(header_row, {"вычеты", "вычет"})
    commission_col = col_index_by_header(header_row, {"комиссия"})

    if cost_col is None:
        cost_col = 2
    if commission_col is None:
        commission_col = 4

    lines: list[ParsedLine] = []
    total = {"budget": Decimal("0.00"), "income": Decimal("0.00"), "cash_status": ""}
    total_found = False

    for row_num, row in rows[2:]:
        name = clean_text(row[name_col])
        key = label_key(name)
        if key.startswith("итого"):
            total = total_row_values(row)
            total_found = True
            break
        if not name:
            continue

        kind = "regular"
        for special_prefix, special_kind in SPECIAL_LABELS.items():
            if key.startswith(special_prefix):
                kind = special_kind
                break
        if kind == "taxes_section":
            continue

        cost = to_decimal(row[cost_col]) if cost_col is not None and cost_col < len(row) else Decimal("0.00")
        fact = to_decimal(row[fact_col]) if fact_col is not None and fact_col < len(row) else Decimal("0.00")
        vat = to_decimal(row[vat_col]) if vat_col is not None and vat_col < len(row) else Decimal("0.00")
        deduction = to_decimal(row[deduction_col]) if deduction_col is not None and deduction_col < len(row) else Decimal("0.00")
        commission = to_decimal(row[commission_col]) if commission_col is not None and commission_col < len(row) else Decimal("0.00")

        lines.append(ParsedLine(row=row_num, name=name, kind=kind, cost=cost, fact=fact, vat=vat, deduction=deduction, commission=commission))

    if not total_found:
        warnings.append(f"{sheet_name} row {start_row_num}: не найдена строка Итого")

    regular_lines = [line for line in lines if line.kind == "regular"]
    agency_lines = [line for line in lines if line.kind == "agency"]
    tax_lines = [line for line in lines if line.kind == "taxes"]
    manager_lines = [line for line in lines if line.kind == "manager_salary"]
    vat_lines = [line for line in lines if line.kind == "vat_line"]

    all_cost_sum = sum((line.cost for line in lines), Decimal("0.00"))
    regular_cost_sum = sum((line.cost for line in regular_lines), Decimal("0.00"))
    fact_sum = sum((line.fact for line in lines), Decimal("0.00"))
    vat_sum = sum((line.vat for line in lines), Decimal("0.00")) + sum((line.cost for line in vat_lines), Decimal("0.00"))
    deduction_sum = sum((line.deduction for line in lines), Decimal("0.00"))
    commission_sum = sum((line.commission for line in lines), Decimal("0.00"))
    agency_commission = sum((line.cost or line.commission for line in agency_lines), Decimal("0.00"))
    taxes_net = sum((abs(line.commission) if line.commission else line.fact - line.deduction for line in tax_lines), Decimal("0.00"))
    manager_salary = sum((abs(line.commission) if line.commission else abs(line.fact) for line in manager_lines), Decimal("0.00"))

    total_budget = total["budget"] or all_cost_sum
    total_income = total["income"] or commission_sum

    return ParsedEvent(
        sheet=sheet_name,
        source_row=start_row_num,
        title=title,
        raw_date=raw_date,
        event_date=event_date,
        old_manager=old_manager,
        total_budget=total_budget,
        total_income=total_income,
        cash_status=clean_text(total.get("cash_status")),
        regular_cost_sum=regular_cost_sum,
        all_cost_sum=all_cost_sum,
        fact_sum=fact_sum,
        vat_sum=vat_sum,
        deduction_sum=deduction_sum,
        commission_sum=commission_sum,
        agency_commission=agency_commission,
        taxes_net=taxes_net,
        manager_salary=manager_salary,
        line_count=len(lines),
        lines_preview=[line.as_dict() for line in lines[:8]],
        lines=lines,
    )


def manager_lookup(db: Session | None, target_manager_name: str) -> dict[str, Any]:
    if db is None:
        return {"requested_name": target_manager_name, "found": None}
    user = db.execute(
        select(User).where(func.lower(User.name) == target_manager_name.strip().lower()).limit(1)
    ).scalar_one_or_none()
    if not user:
        return {"requested_name": target_manager_name, "found": False}
    return {
        "requested_name": target_manager_name,
        "found": True,
        "id": user.id,
        "name": user.name,
        "role": user.role,
        "department_id": user.department_id,
        "is_active": bool(user.is_active),
    }




def parse_legacy_events_2026_xlsx(raw_xlsx: bytes) -> tuple[list[ParsedEvent], list[dict[str, Any]], dict[str, Any], list[str]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Add openpyxl to requirements.txt and redeploy.")

    warnings: list[str] = []
    workbook = load_workbook(BytesIO(raw_xlsx), data_only=True, read_only=True)
    sheet_names = set(workbook.sheetnames)
    missing = [sheet for sheet in TARGET_SHEETS if sheet not in sheet_names]
    if missing:
        raise ValueError(f"В файле нет нужных листов: {', '.join(missing)}")

    events: list[ParsedEvent] = []
    sheet_summaries: list[dict[str, Any]] = []

    for sheet_name, month in TARGET_SHEETS.items():
        sheet = workbook[sheet_name]
        materialized_rows: list[tuple[int, list[Any]]] = []
        for row_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            values = row_values(row, 16)
            if any(value is not None and clean_text(value) != "" for value in values):
                materialized_rows.append((row_num, values))

        starts: list[int] = []
        for idx, (_, values) in enumerate(materialized_rows):
            if is_event_start(values):
                starts.append(idx)
        starts.append(len(materialized_rows))

        sheet_events: list[ParsedEvent] = []
        for start_idx, next_idx in zip(starts, starts[1:]):
            block = materialized_rows[start_idx:next_idx]
            parsed = parse_event_block(sheet_name, month, block, warnings)
            if parsed:
                sheet_events.append(parsed)
                events.append(parsed)

        sheet_summaries.append(
            {
                "sheet": sheet_name,
                "events": len(sheet_events),
                "total_budget": money_out(sum((event.total_budget for event in sheet_events), Decimal("0.00"))),
                "total_income": money_out(sum((event.total_income for event in sheet_events), Decimal("0.00"))),
                "regular_cost_sum": money_out(sum((event.regular_cost_sum for event in sheet_events), Decimal("0.00"))),
                "fact_sum": money_out(sum((event.fact_sum for event in sheet_events), Decimal("0.00"))),
                "agency_commission": money_out(sum((event.agency_commission for event in sheet_events), Decimal("0.00"))),
                "taxes_net": money_out(sum((event.taxes_net for event in sheet_events), Decimal("0.00"))),
                "manager_salary": money_out(sum((event.manager_salary for event in sheet_events), Decimal("0.00"))),
            }
        )

    totals = {
        "events": len(events),
        "total_budget": money_out(sum((event.total_budget for event in events), Decimal("0.00"))),
        "total_income": money_out(sum((event.total_income for event in events), Decimal("0.00"))),
        "regular_cost_sum": money_out(sum((event.regular_cost_sum for event in events), Decimal("0.00"))),
        "fact_sum": money_out(sum((event.fact_sum for event in events), Decimal("0.00"))),
        "vat_sum": money_out(sum((event.vat_sum for event in events), Decimal("0.00"))),
        "deduction_sum": money_out(sum((event.deduction_sum for event in events), Decimal("0.00"))),
        "agency_commission": money_out(sum((event.agency_commission for event in events), Decimal("0.00"))),
        "taxes_net": money_out(sum((event.taxes_net for event in events), Decimal("0.00"))),
        "manager_salary": money_out(sum((event.manager_salary for event in events), Decimal("0.00"))),
    }
    return events, sheet_summaries, totals, warnings


def legacy_event_id_for(event: ParsedEvent) -> str:
    month = TARGET_SHEETS.get(event.sheet, 0)
    return f"legacy-2026-{month:02d}-row-{event.source_row}"


def legacy_item_key_for(event: ParsedEvent, line: ParsedLine, idx: int) -> str:
    return f"{legacy_event_id_for(event)}-item-{idx:03d}-row-{line.row}"


def dry_run_legacy_events_2026_xlsx(raw_xlsx: bytes, *, target_manager_name: str = "Тест", db: Session | None = None) -> dict[str, Any]:
    if load_workbook is None:
        return {
            "ok": False,
            "error": "openpyxl is not installed. Add openpyxl to requirements.txt and redeploy.",
        }
    try:
        events, sheet_summaries, totals, warnings = parse_legacy_events_2026_xlsx(raw_xlsx)
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}

    existing_legacy_ids: set[str] = set()
    if db is not None and events:
        ids = [legacy_event_id_for(event) for event in events]
        existing_legacy_ids = set(db.execute(select(Event.legacy_event_id).where(Event.legacy_event_id.in_(ids))).scalars().all())

    return {
        "ok": True,
        "dry_run": True,
        "source": "legacy_google_sheets_xlsx",
        "period": "2026-01..2026-04",
        "target_manager": manager_lookup(db, target_manager_name),
        "sheets": sheet_summaries,
        "totals": totals,
        "events": [event.as_dict() | {"legacy_event_id": legacy_event_id_for(event), "already_imported": legacy_event_id_for(event) in existing_legacy_ids} for event in events],
        "existing_events": len(existing_legacy_ids),
        "warnings": warnings,
        "notes": [
            "Импорт оплат/заявок/Telegram не выполняется.",
            "Для дат год и месяц берутся из названия листа; день берётся из ячейки даты.",
            "Это dry-run: база данных не меняется.",
            "Боевой импорт пропускает уже загруженные legacy_event_id, чтобы не создать дубли.",
        ],
    }


def event_line_item_type(line: ParsedLine) -> str:
    key = label_key(line.name)
    if line.kind == "manager_salary" or ("менедж" in key and ("зп" in key or "%" in key)):
        return "manager_salary"
    if key == "координатор" or line.kind == "coordinator":
        return "coordinator"
    return "regular"


def import_legacy_events_2026_xlsx(raw_xlsx: bytes, *, target_manager_name: str = "Тест", db: Session, imported_by: str = "legacy_page") -> dict[str, Any]:
    if load_workbook is None:
        return {"ok": False, "error": "openpyxl is not installed. Add openpyxl to requirements.txt and redeploy."}
    manager = db.execute(select(User).where(func.lower(User.name) == target_manager_name.strip().lower()).limit(1)).scalar_one_or_none()
    if not manager:
        return {"ok": False, "error": f"Менеджер '{target_manager_name}' не найден"}
    if not manager.department_id:
        return {"ok": False, "error": f"У менеджера '{manager.name}' не указан отдел"}

    try:
        events, sheet_summaries, totals, warnings = parse_legacy_events_2026_xlsx(raw_xlsx)
    except ValueError as exc:
        return {"ok": False, "error": str(exc)}

    stats = {
        "events_found": len(events),
        "events_created": 0,
        "events_skipped_existing": 0,
        "items_created": 0,
        "payment_requests_created": 0,
    }
    imported_events: list[dict[str, Any]] = []

    try:
        for parsed in events:
            legacy_id = legacy_event_id_for(parsed)
            existing = db.execute(select(Event).where(Event.legacy_event_id == legacy_id).limit(1)).scalar_one_or_none()
            if existing:
                stats["events_skipped_existing"] += 1
                imported_events.append({
                    "legacy_event_id": legacy_id,
                    "title": parsed.title,
                    "event_date": parsed.event_date,
                    "status": "skipped_existing",
                    "event_id": existing.id,
                })
                continue

            event = Event(
                legacy_event_id=legacy_id,
                client_name=parsed.title or "Архив 2026",
                title=parsed.title or "Архивное мероприятие 2026",
                event_date=date.fromisoformat(parsed.event_date),
                department_id=manager.department_id,
                manager_id=manager.id,
                status="draft",
                money_status="waiting_money",
                client_calc_type="cash",
                manager_percent=Decimal("21.00"),
                agency_commission_amount=Decimal("0.00"),
                customer_paid_amount=parsed.total_budget,
                agency_commission_spread_enabled="false",
                simplified_bank_tax_percent=None,
            )
            db.add(event)
            db.flush()
            stats["events_created"] += 1

            if parsed.lines:
                source_lines = parsed.lines
            else:
                source_lines = [
                    ParsedLine(
                        row=parsed.source_row,
                        name="Бюджет по старой таблице",
                        kind="regular",
                        cost=parsed.total_budget,
                        fact=max(Decimal("0.00"), parsed.total_budget - parsed.total_income),
                    )
                ]

            created_item_index = 0
            for idx, line in enumerate(source_lines, start=1):
                if not line.name:
                    continue
                if line.cost == 0 and line.fact == 0 and line.vat == 0 and line.deduction == 0 and line.commission == 0:
                    continue
                created_item_index += 1
                item_type = event_line_item_type(line)
                fact_value = line.fact if line.fact != 0 else Decimal("0.00")
                item = EventItem(
                    legacy_item_key=legacy_item_key_for(parsed, line, created_item_index),
                    event_id=event.id,
                    item_type=item_type,
                    external_name=line.name[:255],
                    external_price=line.cost,
                    external_quantity=Decimal("1.00"),
                    external_days=Decimal("1.00"),
                    external_amount=line.cost,
                    external_note=f"Импорт январь–апрель 2026: {parsed.sheet}, строка {line.row}; старый менеджер: {parsed.old_manager or '—'}",
                    amount_fact=fact_value,
                    paid_amount=Decimal("0.00"),
                    payment_method="invoice" if (line.vat or line.deduction) else "cash",
                    iin_bin=None,
                    iin_bin_locked=False,
                    tax_check_status=None,
                    vat_amount=line.vat,
                    deduction_amount=line.deduction,
                    internal_note=(
                        f"legacy_source=google_sheets_2026_jan_apr; legacy_event={legacy_id}; "
                        f"kind={line.kind}; old_manager={parsed.old_manager or '—'}; "
                        f"legacy_commission={money_out(line.commission)}"
                    ),
                    sort_order=created_item_index,
                    is_deleted=False,
                )
                db.add(item)
                stats["items_created"] += 1

            imported_events.append({
                "legacy_event_id": legacy_id,
                "title": parsed.title,
                "event_date": parsed.event_date,
                "event_id": event.id,
                "items": len(source_lines),
                "status": "created",
            })

        db.commit()
    except Exception:
        db.rollback()
        raise

    return {
        "ok": True,
        "dry_run": False,
        "source": "legacy_google_sheets_xlsx",
        "period": "2026-01..2026-04",
        "target_manager": manager_lookup(db, target_manager_name),
        "sheets": sheet_summaries,
        "totals": totals,
        "stats": stats,
        "events": imported_events,
        "warnings": warnings,
        "notes": [
            "Созданы только мероприятия и строки сметы.",
            "Оплаты, заявки и Telegram не создавались.",
            "Повторный запуск пропускает уже импортированные legacy_event_id.",
            f"Импортировано пользователем/источником: {imported_by}.",
        ],
    }
